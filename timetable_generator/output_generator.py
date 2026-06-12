"""
Timetable Output Generator

Generates timetable outputs in various formats:
- PDF (using ReportLab)
- HTML (responsive, printable)
- Excel (using openpyxl)

Matches the exact CUI timetable format:
- Header: "COMSATS Vehari Centralized Timetable (V-2)-Spring-2026"
- Slot columns with exact times
- Break column between slot 3 and 4
- Rotated day labels
- Merged cells for labs
"""
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from collections import defaultdict
import html
import sys

sys.path.insert(0, 'c:/Users/Fareed Bhatti/Desktop/CUI Campus bot 1')

from timetable_generator.cpsat_generator import ScheduledSession, GeneratorResult

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch, cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, 
        Spacer, PageBreak
    )
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Constants
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
SLOT_TIMES = [
    ("8:30", "10:00", "AM"),
    ("10:00", "11:30", "AM"),
    ("11:30", "1:00", "PM"),
    ("1:30", "3:00", "PM"),
    ("3:00", "4:30", "PM"),
    ("4:30", "6:00", "PM"),
]


class TimetableOutputGenerator:
    """Generates timetable outputs in PDF, HTML, and Excel formats."""
    
    def __init__(self, output_dir: str = "output/timetables"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.semester = "Spring-2026"
    
    def generate_all(
        self, 
        result: GeneratorResult,
        formats: List[str] = ["html", "pdf"]
    ) -> Dict[str, List[str]]:
        """
        Generate timetables for all batches in specified formats.
        
        Args:
            result: GeneratorResult from the solver
            formats: List of formats to generate ("html", "pdf", "xlsx")
            
        Returns:
            Dict mapping format to list of generated file paths
        """
        generated_files: Dict[str, List[str]] = defaultdict(list)
        
        for batch, sessions in result.schedule_by_batch.items():
            if "html" in formats:
                html_path = self.generate_html(batch, sessions)
                generated_files["html"].append(html_path)
            
            if "pdf" in formats and REPORTLAB_AVAILABLE:
                pdf_path = self.generate_pdf(batch, sessions)
                generated_files["pdf"].append(pdf_path)
        
        # Generate master schedule Excel
        if "xlsx" in formats and OPENPYXL_AVAILABLE:
            xlsx_path = self.generate_master_excel(result)
            generated_files["xlsx"].append(xlsx_path)
        
        return generated_files
    
    def _build_grid(
        self, 
        sessions: List[ScheduledSession]
    ) -> Dict[Tuple[int, int], ScheduledSession]:
        """
        Build a grid mapping (day, slot) to session.
        For labs, only map the starting slot.
        """
        grid: Dict[Tuple[int, int], ScheduledSession] = {}
        
        for session in sessions:
            key = (session.day_idx, session.slot_start)
            grid[key] = session
        
        return grid
    
    def _get_skip_slots(
        self, 
        sessions: List[ScheduledSession]
    ) -> set:
        """Get slots that should be skipped (second slot of labs)."""
        skip = set()
        for s in sessions:
            if s.slot_span == 2:
                skip.add((s.day_idx, s.slot_end))
        return skip
    
    # =========================================================================
    # HTML Generation
    # =========================================================================
    
    def generate_html(
        self, 
        batch: str, 
        sessions: List[ScheduledSession]
    ) -> str:
        """Generate HTML timetable for a batch."""
        grid = self._build_grid(sessions)
        skip_slots = self._get_skip_slots(sessions)
        
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Timetable - {html.escape(batch)}</title>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        body {{
            font-family: Arial, sans-serif;
            background: #f5f5f5;
            padding: 20px;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 20px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid #2c5282;
        }}
        .title {{
            font-size: 16px;
            color: #333;
            text-decoration: underline;
            margin-bottom: 10px;
        }}
        .batch-name {{
            font-size: 28px;
            font-weight: bold;
            color: #2c5282;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }}
        th, td {{
            border: 1px solid #666;
            padding: 8px 4px;
            text-align: center;
            vertical-align: middle;
        }}
        th {{
            background: #2c5282;
            color: white;
            font-size: 11px;
            height: 60px;
        }}
        th.day-col {{
            width: 70px;
            background: #d4d4d4;
            color: #333;
        }}
        th.break-col {{
            width: 80px;
            background: #e8e8e8;
            color: #333;
        }}
        td.day-cell {{
            background: #d4d4d4;
            font-weight: bold;
            font-size: 12px;
            height: 90px;
            writing-mode: vertical-rl;
            text-orientation: mixed;
            transform: rotate(180deg);
        }}
        td.break-cell {{
            background: #e8e8e8;
            color: #666;
            writing-mode: vertical-rl;
            text-orientation: mixed;
            transform: rotate(180deg);
            font-weight: bold;
        }}
        td.slot-cell {{
            height: 90px;
            background: #f8f9ff;
            font-size: 10px;
        }}
        td.slot-cell.lab {{
            background: #fff0e6;
        }}
        td.slot-cell.empty {{
            background: #fafafa;
        }}
        .course-name {{
            font-weight: bold;
            font-size: 11px;
            color: #333;
            margin-bottom: 4px;
        }}
        .room-code {{
            font-size: 10px;
            color: #555;
            margin-bottom: 2px;
        }}
        .teacher-name {{
            font-size: 9px;
            color: #666;
        }}
        .slot-time {{
            font-size: 9px;
            font-weight: normal;
            display: block;
            margin-top: 4px;
        }}
        .footer {{
            margin-top: 20px;
            display: flex;
            justify-content: space-between;
            font-size: 10px;
            color: #666;
            padding-top: 10px;
            border-top: 1px solid #ddd;
        }}
        @media print {{
            body {{ background: white; padding: 0; }}
            .container {{ box-shadow: none; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="title">COMSATS Vehari Centralized Timetable (V-2)-{self.semester}</div>
            <div class="batch-name">{html.escape(batch)}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th class="day-col">Days</th>
"""
        
        # Add slot headers
        for i, (start, end, period) in enumerate(SLOT_TIMES):
            slot_num = i + 1 if i < 3 else i  # Account for break after slot 3
            html_content += f"""                    <th>Slot {slot_num}<span class="slot-time">{start}-{end} {period}</span></th>
"""
            # Add break column after slot 3
            if i == 2:
                html_content += """                    <th class="break-col">Break<span class="slot-time">1:00-1:30 PM</span></th>
"""
        
        html_content += """                </tr>
            </thead>
            <tbody>
"""
        
        # Add rows for each day
        for day_idx, day_name in enumerate(DAY_NAMES):
            html_content += f"""                <tr>
                    <td class="day-cell">{day_name}</td>
"""
            
            # Add cells for each slot
            for slot in range(1, 7):
                # Add break column after slot 3
                if slot == 4:
                    html_content += """                    <td class="break-cell">2 Hrs / Break</td>
"""
                
                key = (day_idx, slot)
                
                # Check if this slot should be skipped (second slot of lab)
                if key in skip_slots:
                    continue
                
                if key in grid:
                    session = grid[key]
                    colspan = session.slot_span
                    cell_class = "slot-cell lab" if session.is_lab else "slot-cell"
                    
                    colspan_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                    
                    course_name = html.escape(session.session.course_name)
                    teacher = html.escape(session.session.teacher_name)
                    room = html.escape(session.room_code) if session.room_code else ""
                    
                    html_content += f"""                    <td class="{cell_class}"{colspan_attr}>
                        <div class="course-name">{course_name}</div>
                        <div class="room-code">{room}</div>
                        <div class="teacher-name">{teacher}</div>
                    </td>
"""
                else:
                    html_content += """                    <td class="slot-cell empty"></td>
"""
            
            html_content += """                </tr>
"""
        
        html_content += f"""            </tbody>
        </table>
        
        <div class="footer">
            <span>COMSATS Centralized Timetable - {datetime.now().strftime('%Y-%m-%d')}</span>
            <span>Generated by Timetable Generator</span>
        </div>
    </div>
</body>
</html>
"""
        
        # Save file
        safe_batch = batch.replace("/", "-").replace("\\", "-")
        file_path = self.output_dir / f"{safe_batch}.html"
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return str(file_path)
    
    # =========================================================================
    # PDF Generation
    # =========================================================================
    
    def generate_pdf(
        self, 
        batch: str, 
        sessions: List[ScheduledSession]
    ) -> str:
        """Generate PDF timetable for a batch."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab not installed. Run: pip install reportlab")
        
        grid = self._build_grid(sessions)
        skip_slots = self._get_skip_slots(sessions)
        
        safe_batch = batch.replace("/", "-").replace("\\", "-")
        file_path = self.output_dir / f"{safe_batch}.pdf"
        
        # Create document
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=landscape(A4),
            leftMargin=0.4*inch,
            rightMargin=0.4*inch,
            topMargin=0.4*inch,
            bottomMargin=0.4*inch
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=TA_CENTER,
            textColor=colors.black,
            spaceAfter=6
        )
        batch_style = ParagraphStyle(
            'Batch',
            parent=styles['Heading2'],
            fontSize=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            spaceAfter=12
        )
        cell_style = ParagraphStyle(
            'Cell',
            parent=styles['Normal'],
            fontSize=7,
            alignment=TA_CENTER,
            leading=9
        )
        
        # Build content
        elements = []
        
        # Header
        elements.append(Paragraph(
            f"<u>COMSATS Vehari Centralized Timetable (V-2)-{self.semester}</u>",
            title_style
        ))
        elements.append(Paragraph(batch, batch_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Build table data
        # Header row
        header_row = ['Days']
        for i, (start, end, period) in enumerate(SLOT_TIMES):
            slot_num = i + 1 if i < 3 else i
            header_row.append(f"Slot {slot_num}\n{start}-{end}\n{period}")
            if i == 2:
                header_row.append("Break\n1:00-1:30\nPM")
        
        # Data rows
        table_data = [header_row]
        
        for day_idx, day_name in enumerate(DAY_NAMES):
            row = [day_name]
            
            col_idx = 1
            for slot in range(1, 7):
                if slot == 4:
                    row.append("2 Hrs\nBreak")
                    col_idx += 1
                
                key = (day_idx, slot)
                
                if key in skip_slots:
                    continue
                
                if key in grid:
                    session = grid[key]
                    cell_text = f"{session.session.course_name}\n"
                    if session.room_code:
                        cell_text += f"{session.room_code}\n"
                    cell_text += session.session.teacher_name
                    row.append(cell_text)
                else:
                    row.append("")
                
                col_idx += 1
            
            table_data.append(row)
        
        # Calculate column widths
        page_width = landscape(A4)[0] - 0.8*inch
        day_col = 0.7*inch
        break_col = 0.8*inch
        remaining = page_width - day_col - break_col
        slot_col = remaining / 6
        
        col_widths = [day_col, slot_col, slot_col, slot_col, break_col, 
                      slot_col, slot_col, slot_col]
        
        # Create table
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Style table
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.17, 0.32, 0.51)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            
            # Day column
            ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.83, 0.83, 0.83)),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            
            # Break column
            ('BACKGROUND', (4, 0), (4, -1), colors.Color(0.91, 0.91, 0.91)),
            
            # Data cells
            ('FONTSIZE', (1, 1), (-1, -1), 7),
            ('BACKGROUND', (1, 1), (3, -1), colors.Color(0.97, 0.97, 1.0)),
            ('BACKGROUND', (5, 1), (7, -1), colors.Color(0.97, 0.97, 1.0)),
        ])
        
        # Highlight lab cells
        for day_idx in range(5):
            for slot in range(1, 7):
                key = (day_idx, slot)
                if key in grid and grid[key].is_lab:
                    row = day_idx + 1
                    col = slot if slot <= 3 else slot + 1
                    style.add('BACKGROUND', (col, row), (col, row), 
                             colors.Color(1.0, 0.94, 0.9))
        
        table.setStyle(style)
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey
        )
        elements.append(Paragraph(
            f"COMSATS Centralized Timetable - {datetime.now().strftime('%Y-%m-%d')}",
            footer_style
        ))
        
        # Build PDF
        doc.build(elements)
        
        return str(file_path)
    
    # =========================================================================
    # Excel Generation
    # =========================================================================
    
    def generate_master_excel(self, result: GeneratorResult) -> str:
        """Generate master schedule Excel file with all batches."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Summary headers
        headers = ["Batch/Section", "Day", "Slot", "Course", "Teacher", 
                   "Room", "Type", "Slot Span"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C5282", end_color="2C5282", 
                                   fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Add all sessions
        row = 2
        for session in result.scheduled_sessions:
            ws_summary.cell(row=row, column=1, value=session.session.batch_section)
            ws_summary.cell(row=row, column=2, value=DAY_NAMES[session.day_idx])
            ws_summary.cell(row=row, column=3, value=session.slot_start)
            ws_summary.cell(row=row, column=4, value=session.session.course_name)
            ws_summary.cell(row=row, column=5, value=session.session.teacher_name)
            ws_summary.cell(row=row, column=6, value=session.room_code)
            ws_summary.cell(row=row, column=7, value="LAB" if session.is_lab else "LEC")
            ws_summary.cell(row=row, column=8, value=session.slot_span)
            row += 1
        
        # Auto-adjust column widths
        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Create individual batch sheets
        for batch in sorted(result.schedule_by_batch.keys()):
            sessions = result.schedule_by_batch[batch]
            safe_name = batch[:30].replace("/", "-")  # Sheet name max 31 chars
            ws = wb.create_sheet(title=safe_name)
            
            self._create_batch_sheet(ws, batch, sessions)
        
        # Save
        file_path = self.output_dir / "master_schedule.xlsx"
        wb.save(str(file_path))
        
        return str(file_path)
    
    def _create_batch_sheet(
        self, 
        ws, 
        batch: str, 
        sessions: List[ScheduledSession]
    ):
        """Create a formatted timetable sheet for a batch."""
        grid = self._build_grid(sessions)
        skip_slots = self._get_skip_slots(sessions)
        
        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = f"COMSATS Vehari Centralized Timetable - {self.semester}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:I2')
        ws['A2'] = batch
        ws['A2'].font = Font(bold=True, size=16)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers (row 4)
        headers = ['Days', 'Slot 1\n8:30-10:00', 'Slot 2\n10:00-11:30', 
                  'Slot 3\n11:30-1:00', 'Break\n1:00-1:30', 
                  'Slot 4\n1:30-3:00', 'Slot 5\n3:00-4:30', 'Slot 6\n4:30-6:00']
        
        header_fill = PatternFill(start_color="2C5282", end_color="2C5282", 
                                  fill_type="solid")
        break_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8",
                                fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF" if col != 5 else "000000")
            cell.fill = header_fill if col != 5 else break_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', 
                                       wrap_text=True)
        
        # Data rows
        for day_idx, day_name in enumerate(DAY_NAMES):
            row = 5 + day_idx
            
            # Day cell
            ws.cell(row=row, column=1, value=day_name)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color="D4D4D4", end_color="D4D4D4", fill_type="solid")
            
            # Break cell
            ws.cell(row=row, column=5, value="Break")
            ws.cell(row=row, column=5).fill = break_fill
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            
            # Slot cells
            for slot in range(1, 7):
                col = slot + 1 if slot <= 3 else slot + 2
                key = (day_idx, slot)
                
                if key in skip_slots:
                    continue
                
                if key in grid:
                    session = grid[key]
                    cell_text = f"{session.session.course_name}\n"
                    if session.room_code:
                        cell_text += f"{session.room_code}\n"
                    cell_text += session.session.teacher_name
                    
                    ws.cell(row=row, column=col, value=cell_text)
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True)
                    
                    if session.is_lab:
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFF0E6", end_color="FFF0E6", 
                            fill_type="solid")
        
        # Set row heights
        ws.row_dimensions[4].height = 50
        for i in range(5, 10):
            ws.row_dimensions[i].height = 80
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        for col in 'BCDEFGHI':
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['E'].width = 10  # Break column


def generate_outputs(
    result: GeneratorResult,
    output_dir: str = "output/timetables",
    formats: List[str] = ["html", "pdf", "xlsx"]
) -> Dict[str, List[str]]:
    """
    Convenience function to generate all outputs.
    
    Args:
        result: GeneratorResult from the solver
        output_dir: Output directory
        formats: List of formats to generate
        
    Returns:
        Dict mapping format to list of generated file paths
    """
    generator = TimetableOutputGenerator(output_dir)
    return generator.generate_all(result, formats)
